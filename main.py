from openpyxl.utils.cell import coordinate_from_string, column_index_from_string
from os import listdir
from os.path import isfile, join
import openpyxl as op
import pathlib
import warnings

def remove_merged_cells(sheet):

    for cell_range in sheet.merged_cells:  # on retire toutes les cellules merged

        range_cell = str(cell_range).split(":")
        range_cell[0] = coordinate_from_string(range_cell[0])
        range_cell[1] = coordinate_from_string(range_cell[1])
        column = column_index_from_string(range_cell[0][0])

        sheet.unmerge_cells(str(cell_range))

        for c in range(range_cell[0][1], range_cell[1][1] + 1):
            sheet.cell(c, column=column).value = sheet.cell(range_cell[0][1], column=column).value

    reference.save("reference.xlsx")  # on enregistre

def load_files():

    files = [f for f in listdir("./") if isfile(join("./", f))]

    for file in files:

        extension = pathlib.Path(file).suffix

        if extension != ".xls" and extension != ".xlsx":
            files.remove(file)

    if "reference.xlsx" in files:
        files.remove("reference.xlsx")

    if "main.py" in files:
        files.remove("main.py")

    return files

def load_reference_file():

    for sheet in reference.worksheets:  # pour chaque onglet

        remove_merged_cells(sheet)

        previous_loc = ""

        for i in range(int(reference_row_index), int(sheet.max_row) + 1):  # pour chaque ligne du premier document

            if(sheet.cell(i, column=int(reference_column_index)).value):

                temp = (sheet.cell(i, column=int(reference_column_index)).value)

                if temp != None and type(temp) != int and type(temp) != float:

                    try:
                        tab_temp = temp.split(";")  # on split selon le ; pour récupérer les différetes cotes
                    except ValueError:
                        continue

                    if "" in tab_temp:
                        tab_temp.remove("")

                    for j in range(0, len(tab_temp)):
                        tab_temp[j] = tab_temp[j].strip()
                        if "E" not in tab_temp[j]:
                            if "W" in tab_temp[j] and "LO" not in tab_temp[j]:
                                tab_temp[j] = tab_temp[j] + "LO"

                    current_loc = sheet.cell(i, column=1).value # clé de la localisation actuelle

                    if current_loc == None:
                        current_loc = previous_loc

                    current_loc = current_loc.strip()

                    if current_loc in storage1.keys():

                        for t in range(0, len(tab_temp)):
                            storage1[current_loc].append(tab_temp[t])  # si la clé existe on ajoute
                    else:
                        storage1[current_loc] = tab_temp  # sinon on crée puis on ajoute

                    previous_loc = current_loc

def compute_beginning_rowAndColumn(sheet):

    column_index = 1
    row_index = 1

    while (sheet.cell(1, column=column_index)).value != "Series":
        column_index += column_index

    while(sheet.cell(row_index, column=1)).value != "Localisation":
        row_index += row_index

    return row_index+1, column_index

def load_excel(sheet):
    
    for i in range(int(doc_row_index), int(sheet.max_row) + 1):

       temp = (sheet.cell(i, column=int(doc_column_index)).value)

       if temp != None:
        tab_temp = temp.split(";")  # on split selon le ; pour récupérer les différetes cotes

        if "" in tab_temp:
            tab_temp.remove("")

        for j in range (0, len(tab_temp)):
            tab_temp[j] = tab_temp[j].strip()

            if "E" not in tab_temp[j]:
                if "W" in tab_temp[j] and "LO" not in tab_temp[j]:
                    tab_temp[j] = tab_temp[j] + "LO"

        current_loc = sheet.cell(i, column=1).value  # localisation courante

        if current_loc == None:
            current_loc = previous_loc

        current_loc = current_loc.strip()
        if current_loc in storage2.keys():

            for t in range(0, len(tab_temp)):
                storage2[current_loc].append(tab_temp[t])  # si la clé existe on ajoute

        else:
            storage2[current_loc] = tab_temp  # sinon on crée puis on ajoute

        previous_loc = current_loc
    

if __name__ == '__main__':

    warnings.filterwarnings('ignore', category=UserWarning, module='openpyxl')
    # variables
    files = load_files()
    print("Chargement des fichiers: " + str(files) + "...")

    storage1 = dict()
    storage2 = dict()

    print("Chargement du fichier de référence...")

    reference = op.load_workbook("reference.xlsx", read_only=False) # document de référence
    reference_column_index = 3  # colonne et ligne de départ du excel de référence
    reference_row_index = 3

    # chargement du fichier de référence
    load_reference_file()

    for r in range(0, len(files)):

        doc = op.load_workbook(files[r])
        sheet = doc.worksheets[0]
        doc_row_index, doc_column_index = compute_beginning_rowAndColumn(sheet)

    # chargement des autres excels
        load_excel(sheet)

    print("Calcul des différences en cours...")

    f = open("resultat.txt", "w+")

    for key in storage1.keys():

        if key in storage2.keys():

            temp = set(storage1[key]) ^ set(storage2[key])
            if "" in temp:
                temp.remove("")

            if temp == set():
                f.write(key + ": Aucun changement\n")

            else:

                f.write(key + ":")
                f.write(str(temp))
                f.write("\n")

    f.close()  # fermeture du fichier
    print("Programme terminé, résultats disponibles dans le fichier resultat.txt")
    input("Veuillez appuyer sur Entrée afin de fermer le programme")